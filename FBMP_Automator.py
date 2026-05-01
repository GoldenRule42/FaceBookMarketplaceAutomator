#!/usr/bin/env python3
"""
Facebook Marketplace Automator
Uses Google's Gemini 2.5 Flash-Lite vision model to analyze product photos
and automatically generate optimized Marketplace listings in an Excel format
ready for bulk upload.

Prerequisites:
- pip install tqdm openpyxl google-genai Pillow
- A Google Gemini API Key (set as GEMINI_API_KEY env variable or placed in api_key.txt)
- A base 'Marketplace_Bulk_Upload_Template.xlsx' in the root directory.
"""

import os
import sys
import json
import time
import glob
import logging
import shutil
from datetime import datetime
from pathlib import Path

# Third-party libraries
try:
    from tqdm import tqdm
    from openpyxl import load_workbook
    from google import genai
    from PIL import Image
except ImportError as e:
    print(f"Error: A required library is not installed. Please run:\n"
          f"pip install tqdm openpyxl google-genai Pillow\n"
          f"Original error: {e}", file=sys.stderr)
    sys.exit(1)

# ───── Configuration ──────────────────────────────────────────────────────────
API_KEY_FILE = "api_key.txt"  # Changed from Establish.txt to a more universal name
INPUT_DIR = "Input"
OUTPUT_BASE_DIR = "Output"
TEMPLATE_XLSX_PATH = "Marketplace_Bulk_Upload_Template.xlsx"

# Auto-create necessary directories for first-time users
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_BASE_DIR, exist_ok=True)

SESSION_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
SESSION_DIR = os.path.join(OUTPUT_BASE_DIR, SESSION_TIMESTAMP)
os.makedirs(SESSION_DIR, exist_ok=True)

LOG_FILE_PATH = os.path.join(SESSION_DIR, "FBMP_Automator_Log.txt")
OUTPUT_XLSX_PATH = os.path.join(SESSION_DIR, f"{SESSION_TIMESTAMP}_bulk_upload.xlsx")
RECOVERY_JSON_PATH = os.path.join(SESSION_DIR, "recovery_backup.json")
SUMMARY_FILE_PATH = os.path.join(SESSION_DIR, f"{SESSION_TIMESTAMP}_final_summary.txt")

# === OPTIMIZED MODEL SELECTION ===
# Model: gemini-2.5-flash-lite
# Reason: The most cost-effective standard for bulk image processing and text extraction.
MODEL_NAME = "gemini-2.5-flash-lite"

# Pricing for gemini-2.5-flash-lite (USD per 1 Million Tokens) - As of mid-2026
INPUT_TOKEN_COST_PER_MILLION = 0.10
OUTPUT_TOKEN_COST_PER_MILLION = 0.40
# =================================================================

# Per Google's best practices, use the File API for any image > 4MB.
INLINE_SIZE_LIMIT_BYTES = 4 * 1024 * 1024 - (100 * 1024) # 3.9 MB to be safe

EXCEL_HEADER_ROW_COUNT = 1
REQUIRED_HEADERS = ["Name", "Price", "Condition", "Description", "Category", "ImageFilename"]


# ───── Helper Functions ───────────────────────────────────────────────────────

def setup_logging(log_file_path: str) -> logging.Logger:
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s [%(levelname)-8s] %(message)s",
        handlers=[
            logging.FileHandler(log_file_path, mode='w', encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    for lib_name in ["google.genai", "google.ai", "httpx", "urllib3", "PIL"]:
        logging.getLogger(lib_name).setLevel(logging.WARNING)
    return logging.getLogger()

def load_api_key(file_path: str) -> str:
    logger = logging.getLogger()
    
    # 1. Check for Environment Variable first (Best Practice)
    env_key = os.environ.get("GEMINI_API_KEY")
    if env_key:
        logger.info("API key loaded successfully from environment variable.")
        return env_key.strip()

    # 2. Fallback to reading from a local text file
    if not os.path.isfile(file_path):
        logger.critical(f"API Key missing. Please set the 'GEMINI_API_KEY' environment variable OR create an '{file_path}' file in the root directory containing only your key.")
        raise FileNotFoundError(f"API key file not found: {file_path}")
    
    with open(file_path, "r", encoding="utf-8") as f:
        api_key = f.read().strip()
    
    if not api_key:
        logger.critical(f"API key file '{file_path}' is empty.")
        raise ValueError("API key file is empty")
    
    logger.info(f"API key loaded successfully from '{file_path}'.")
    return api_key

def sanitize_filename(name: str) -> str:
    illegal_chars = '<>:"/\\|?*'
    for char in illegal_chars:
        name = name.replace(char, '')
    sanitized = " ".join(name.split()).strip()
    return sanitized or "unnamed_product"

def get_prompt_instructions() -> str:
    return (
        "You are an expert at creating product listings for Facebook Marketplace. "
        "Analyze the provided image of a single product. Your task is to extract "
        "the following information and return it as a single, minified JSON object with no "
        "markdown formatting (e.g., ```json ... ```). \n\n"
        "1.  `Name`: A concise and descriptive title for the product.\n"
        "2.  `Price`: An integer or float representing the price in USD. Do not include the '$' symbol.\n"
        "3.  `Condition`: One of the following values only: 'New', 'Used - Like New', 'Used - Good', 'Used - Fair'.\n"
        "4.  `Description`: A brief, bulleted-point-style description of the key features. Use '\\n' for new lines.\n"
        "5.  `Category`: A suitable category for the product (e.g., 'Electronics', 'Furniture', 'Clothing').\n\n"
        "Respond ONLY with the JSON object."
    )

def prepare_workbook(template_path: str, output_path: str, logger: logging.Logger):
    if not os.path.exists(template_path):
        logger.critical(f"Missing required file: '{template_path}'. Please ensure the blank template is in the same folder as this script.")
        raise FileNotFoundError(f"Template not found: {template_path}")
    shutil.copy(template_path, output_path)
    logger.info(f"Copied template to session workbook: {output_path}")
    wb = load_workbook(output_path)
    sheet = wb.active
    for i, header_text in enumerate(REQUIRED_HEADERS, 1):
        if sheet.cell(row=1, column=i).value != header_text:
            sheet.cell(row=1, column=i, value=header_text)
            logger.warning(f"Header in column {i} was missing/incorrect. Set to '{header_text}'.")
    wb.save(output_path)
    return wb, sheet

def create_final_summary(
    start_time, total_discovered, already_done, successes,
    total_input_tokens, total_output_tokens,
    failed_path=None
) -> str:
    end_time = time.monotonic()
    elapsed_seconds = end_time - start_time
    input_cost = (total_input_tokens / 1_000_000) * INPUT_TOKEN_COST_PER_MILLION
    output_cost = (total_output_tokens / 1_000_000) * OUTPUT_TOKEN_COST_PER_MILLION
    total_cost = input_cost + output_cost
    avg_time_per_image = elapsed_seconds / successes if successes > 0 else 0
    elapsed_str = time.strftime("%H:%M:%S", time.gmtime(elapsed_seconds))
    eta_str = time.strftime("%H:%M:%S", time.gmtime(avg_time_per_image * (total_discovered - already_done - successes))) if not failed_path else "00:00:00"

    summary_lines = [
        "════════════════════════════════════════════",
        "         FBMP Automator Final Summary",
        "════════════════════════════════════════════",
        f"Session Timestamp:      {SESSION_TIMESTAMP}",
        f"Status:                 {'HALTED DUE TO ERROR' if failed_path else 'Completed'}",
    ]
    if failed_path:
        summary_lines.append(f"Failing Image:          {os.path.basename(failed_path)}")
    summary_lines.extend([
        "", "--- Processing Stats ---", f"Total Images Discovered:  {total_discovered}",
        f"Already Processed (Skip): {already_done}", f"Successfully Processed:   {successes}",
        "", "--- Token & Cost Stats ---", f"Input Tokens:             {total_input_tokens:,}",
        f"Output Tokens:            {total_output_tokens:,}", f"Total Tokens:             {total_input_tokens + total_output_tokens:,}",
        f"Estimated Cost:           ${total_cost:.4f} USD", "", "--- Timing ---",
        f"Total Elapsed Time:       {elapsed_str}", f"Average Time / Image:     {avg_time_per_image:.2f}s",
        f"Estimated Time Remaining: {eta_str}", "════════════════════════════════════════════",
    ])
    return "\n".join(summary_lines)


# ───── Main Execution ─────────────────────────────────────────────────────────

def main():
    logger = setup_logging(LOG_FILE_PATH)
    logger.info(f"--- Starting FBMP Automator Session: {SESSION_TIMESTAMP} ---")

    all_image_paths, rows_already_done, successful_runs, total_input_tokens, total_output_tokens = [], 0, 0, 0, 0
    start_time = time.monotonic()
    failed_image_path, image_path = None, None
    
    try:
        api_key = load_api_key(API_KEY_FILE)
        client = genai.Client(api_key=api_key)
        
        wb, sheet = prepare_workbook(TEMPLATE_XLSX_PATH, OUTPUT_XLSX_PATH, logger)
        
        # Ensure rows_already_done is accurate by checking populated rows
        rows_already_done = max(0, sheet.max_row - EXCEL_HEADER_ROW_COUNT)
        
        all_image_paths = sorted(glob.glob(os.path.join(INPUT_DIR, "*.*")))
        images_to_process = all_image_paths[rows_already_done:]

        logger.info(f"{len(all_image_paths)} total images found in '{INPUT_DIR}'.")
        if rows_already_done > 0: logger.info(f"Skipping {rows_already_done} images already processed. (Ensure your template is blank below row 1 if this is incorrect).")
        logger.info(f"{len(images_to_process)} new images to process this run.")
        
        if len(images_to_process) == 0:
            logger.info("No images left to process. Exiting cleanly.")
            return

        recovery_records = []
        pbar = tqdm(images_to_process, desc="Processing Images", unit="img")

        for image_path in pbar:
            current_filename = os.path.basename(image_path)
            pbar.set_postfix_str(f"Curr: {current_filename}")
            logger.debug(f"Processing image: {current_filename}")

            file_size = os.path.getsize(image_path)
            prompt_instructions = get_prompt_instructions()
            
            uploaded_file, image_part = None, None
            try:
                if file_size <= INLINE_SIZE_LIMIT_BYTES:
                    logger.debug(f"Image size ({file_size / 1e6:.2f}MB) is within 4MB limit. Using inline method.")
                    image_part = Image.open(image_path)
                    contents = [prompt_instructions, image_part]
                else:
                    logger.info(f"Image size ({file_size / 1e6:.2f}MB) > 4MB. Using File API...")
                    uploaded_file = client.files.upload(file=image_path)
                    logger.info(f"File '{uploaded_file.name}' uploaded.")
                    contents = [prompt_instructions, uploaded_file]

                # --- Retry Logic for High Demand/Rate Limits ---
                max_retries = 4
                retry_delay = 5  # Start with a 5-second wait
                
                for attempt in range(max_retries):
                    try:
                        response = client.models.generate_content(
                            model=MODEL_NAME, 
                            contents=contents
                        )
                        break  # Success! Break out of the retry loop.
                    
                    except Exception as api_err:
                        err_str = str(api_err).lower()
                        # Check if it's a capacity or rate limit issue
                        if "high demand" in err_str or "503" in err_str or "429" in err_str:
                            if "quota" in err_str and "exceeded" in err_str:
                                logger.error("Daily Quota Exceeded. You must wait 24 hours or upgrade to a paid tier.")
                                raise # Don't retry quota errors
                            elif attempt < max_retries - 1:
                                logger.warning(f"API overloaded or rate limited. Retrying in {retry_delay}s... (Attempt {attempt + 1}/{max_retries})")
                                time.sleep(retry_delay)
                                retry_delay *= 2  # Exponential backoff (5s, 10s, 20s...)
                            else:
                                logger.error("Max retries reached. API is persistently overloaded.")
                                raise
                        else:
                            raise

            finally:
                if image_part: image_part.close()
                if uploaded_file:
                    logger.info(f"Deleting uploaded file '{uploaded_file.name}'.")
                    client.files.delete(name=uploaded_file.name)

            cleaned_text = response.text.strip().removeprefix("```json").removesuffix("```").strip()
            parsed_data = json.loads(cleaned_text)
            
            name, price = parsed_data.get("Name", "Unnamed"), parsed_data.get("Price", 0)
            cond, desc = parsed_data.get("Condition", "Used - Good"), parsed_data.get("Description", "")
            cat = parsed_data.get("Category", "Misc.")
            
            new_filename = f"{sanitize_filename(name)}{Path(current_filename).suffix}"
            sheet.append([name, price, cond, desc, cat, new_filename])
            wb.save(OUTPUT_XLSX_PATH)
            logger.info(f"Appended to Excel: '{name}'")
            
            shutil.move(image_path, os.path.join(SESSION_DIR, new_filename))
            logger.info(f"Moved '{current_filename}' to session folder.")
            
            record = {k:v for k,v in zip(REQUIRED_HEADERS, [name, price, cond, desc, cat, new_filename])}
            recovery_records.append(record)
            with open(RECOVERY_JSON_PATH, 'w', encoding='utf-8') as f:
                json.dump(recovery_records, f, indent=2)
            
            if response.usage_metadata:
                total_input_tokens += response.usage_metadata.prompt_token_count
                total_output_tokens += response.usage_metadata.candidates_token_count
            successful_runs += 1

    except Exception as e:
        error_msg = str(e).lower()
        if "503" in error_msg or "unavailable" in error_msg or "timeout" in error_msg:
            failed_image_path = image_path
            logger.critical("="*60)
            logger.critical("  NETWORK CONNECTION FAILED: The script could not reach Google's API.")
            logger.critical("  This is almost always caused by a FIREWALL or ANTIVIRUS program.")
            logger.critical("  Please check all security software and grant python access.")
            logger.critical("="*60)
            logger.debug(f"Full traceback for network error:", exc_info=True)
        else:
            if image_path:
                failed_image_path = image_path
                logger.critical(f"--- HALTING: Unrecoverable error on image '{os.path.basename(failed_image_path)}' ---", exc_info=True)
            else:
                logger.critical("--- HALTING: Unrecoverable error during script initialization ---", exc_info=True)
    
    finally:
        if 'pbar' in locals(): pbar.close()
        logger.info("--- Generating Final Summary ---")
        final_summary = create_final_summary(
            start_time, len(all_image_paths), rows_already_done, successful_runs,
            total_input_tokens, total_output_tokens, failed_image_path
        )
        with open(SUMMARY_FILE_PATH, 'w', encoding='utf-8') as f: f.write(final_summary)
        print("\n" + final_summary)
        logger.info(f"Final summary saved to {SUMMARY_FILE_PATH}")
        logger.info("--- FBMP Automator Session Finished ---")

if __name__ == "__main__":
    main()